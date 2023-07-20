import openpyxl, datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook

def convert_to_excel(action_id,emails,url):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'].value = str(url)
    sheet['A1'].fill = PatternFill(start_color="A8DADC", fill_type="solid")
    i = 2
    for email in emails:
        sheet['A'+str(i)].value = email
        i+=1
    file_name = url
    file_name = file_name.replace('https://','')
    file_name = file_name.replace('http://','')
    file_name = file_name[:file_name.find('/')]
    workbook.save('results/'+file_name+str(action_id)+'.xlsx')
    return file_name+str(action_id)+'.xlsx'

def convert_bulk_to_excel(action_id,emails,urls):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    column = "A"
    i = 0
    for url in urls:
        line = 2
        if len(emails[i]) == 0:
            continue
        sheet[column + '1'].value = str(url)
        sheet[column + '1'].fill = PatternFill(start_color="A8DADC", fill_type="solid")
        for email in emails[i]:
            sheet[str(column)+str(line)].value = email
            line+=1
        column = get_column_letter(column_index_from_string(column) + 1)
        i+=1
    file_name = urls[0]
    file_name = file_name.replace('https://','')
    file_name = file_name.replace('http://','')
    file_name = file_name[:file_name.find('/')]
    workbook.save('results/'+file_name+str(action_id)+'.xlsx')
    return file_name+str(action_id)+'.xlsx'

def update_excel(filename,emails,urls):
    workbook = load_workbook(filename='results/'+filename)
    sheet = workbook.active
    column = "A"
    while sheet[column + '1'].value is not None:
        url = sheet[column + '1'].value
        if url in urls:
            i = urls.index(url)
            del urls[i]
            line = 2
            for email in emails[i]:
                sheet[str(column)+str(line)].value = email
                line+=1
            del emails[i]
        column = get_column_letter(column_index_from_string(column) + 1)
    i=0
    for url in urls:
        line = 2
        if len(emails[i]) == 0:
            continue
        sheet[column + '1'].value = str(url)
        sheet[column + '1'].fill = PatternFill(start_color="A8DADC", fill_type="solid")
        for email in emails[i]:
            sheet[str(column)+str(line)].value = email
            line+=1
        column = get_column_letter(column_index_from_string(column) + 1)
        i+=1
    workbook.save('results/'+filename)